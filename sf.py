import requests
import pandas as pd
import time
import asyncio
from datetime import datetime
import os

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from binance.client import Client

# ================= BINANCE STUFF =================

load_dotenv(".env.local")

API_KEY = os.environ.get("API_KEY")
API_SECRET = os.environ.get("API_SECRET")

client = None

if API_KEY and API_SECRET:
    try:
        client = Client(API_KEY, API_SECRET)
        account = client.get_account()
        print("[spotfutures] Connected. Can trade:", account["canTrade"])
    except Exception as error:
        print("[spotfutures] Binance auth failed, continuing in read-only mode:", error)
else:
    print("[spotfutures] API_KEY/API_SECRET not set, continuing in read-only mode")

# ================= SETTINGS =================
FEE = 0.001
TOTAL_FEES = FEE * 2

THRESHOLD = 0.3  # %
LIVE_FILE = "live_sf.xlsx"
LOG_FILE = "log_sf.xlsx"
TABLE_NAME = "ArbTable"

# Spread filter threshold (0.15%)
MAX_SPREAD = 0.0015

VALID_PAIRS1 = [
    "BTCUSDT", "ETHUSDT", "BNBUSDT", "SOLUSDT",
    "XRPUSDT", "TRXUSDT", "HYPEUSDT", "DOGEUSDT",
    "ZECUSDT"
]


# ================= WEBSOCKET BROADCAST =================
# Mirrors scanner.py's pattern so both engines can share one FastAPI app.

clients = []
main_loop = None


async def broadcast_opportunity(payload):
    dead_clients = []

    for ws in clients:
        try:
            await ws.send_json(payload)
        except Exception:
            dead_clients.append(ws)

    for ws in dead_clients:
        if ws in clients:
            clients.remove(ws)


def send_opportunity(payload):
    if main_loop is None:
        return

    asyncio.run_coroutine_threadsafe(
        broadcast_opportunity(payload),
        main_loop
    )


# ================= API =================
def get_spot_prices():
    url = "https://api.binance.com/api/v3/ticker/bookTicker"
    data = requests.get(url, timeout=5).json()

    return {
        item['symbol']: {
            "bid": float(item['bidPrice']),
            "ask": float(item['askPrice'])
        } for item in data
    }


def get_futures_prices():
    url = "https://fapi.binance.com/fapi/v1/ticker/bookTicker"
    data = requests.get(url, timeout=5).json()

    return {
        item['symbol']: {
            "bid": float(item['bidPrice']),
            "ask": float(item['askPrice'])
        } for item in data
    }


def get_funding_rates():
    url = "https://fapi.binance.com/fapi/v1/premiumIndex"
    data = requests.get(url, timeout=5).json()

    return {
        item['symbol']: {
            "rate": float(item['lastFundingRate']),
            "next_time": item['nextFundingTime'],
        } for item in data
    }


# ================= SPREAD FILTER =================
def is_spread_ok(data, symbol, max_spread=MAX_SPREAD):
    bid = data[symbol]['bid']
    ask = data[symbol]['ask']

    if ask == 0:
        return False

    spread = (ask - bid) / ask
    return spread < max_spread


# ================= LOGIC =================

# Live spread for every tracked pair with acceptable market quality, whether
# or not it currently clears the profit threshold. Genuine cash-and-carry
# opportunities (net_percent > 0) are rare, so a dashboard fed only the
# profitable subset would sit silent almost all the time — this gives it a
# continuous feed to render (and to build the spread chart from later).
def get_all_spreads(spot, futures, funding=None):
    funding = funding or {}
    rows = []

    for symbol in VALID_PAIRS1:
        if symbol not in spot or symbol not in futures:
            continue

        try:
            if not is_spread_ok(spot, symbol):
                continue

            if not is_spread_ok(futures, symbol):
                continue

            spot_price = spot[symbol]['ask']
            fut_price = futures[symbol]['bid']

            spread = ((fut_price - spot_price) / spot_price) * 100
            net = spread - (TOTAL_FEES * 100)

            fr = funding.get(symbol, {})
            funding_rate = fr.get("rate", 0.0)

            rows.append({
                "time": datetime.now().strftime("%H:%M:%S"),
                "symbol": symbol,
                "spot": spot_price,
                "futures": fut_price,
                "spread_percent": round(spread, 4),
                "net_percent": round(net, 4),
                "is_opportunity": net > 0,
                "funding_rate_percent": round(funding_rate * 100, 4),
                # Perpetuals settle funding 3x/day (~1095x/year) — this is the
                # naive extrapolation CMC/exchanges show as "est. APY".
                "funding_apy_percent": round(funding_rate * 100 * 3 * 365, 2),
                "next_funding_time": fr.get("next_time"),
            })

        except Exception:
            continue

    return sorted(rows, key=lambda x: -x["net_percent"])


def find_opportunities(spot, futures, funding=None):
    return [row for row in get_all_spreads(spot, futures, funding) if row["is_opportunity"]]


# ================= EXCEL =================

def update_live(df):
    try:
        if df.empty:
            df = pd.DataFrame([{
                "Time": datetime.now().strftime("%H:%M:%S"),
                "Symbol": "None",
                "Spot": 0,
                "Futures": 0,
                "Spread (%)": 0,
                "Net (%)": 0
            }])

        df.head(15).to_excel(LIVE_FILE, index=False)

    except Exception:
        pass


def append_log_table(df):
    if df.empty:
        return

    if not os.path.exists(LOG_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Log"

        headers = list(df.columns)
        ws.append(headers)

        for row in df.values:
            ws.append(list(row))

        col_letter = get_column_letter(ws.max_column)

        table = Table(
            displayName=TABLE_NAME,
            ref=f"A1:{col_letter}{ws.max_row}"
        )

        style = TableStyleInfo(
            name="TableStyleMedium9",
            showRowStripes=True
        )

        table.tableStyleInfo = style
        ws.add_table(table)

        wb.save(LOG_FILE)

    else:
        wb = load_workbook(LOG_FILE)
        ws = wb["Log"]

        for row in df.values:
            ws.append(list(row))

        table = ws.tables[TABLE_NAME]

        col_letter = get_column_letter(ws.max_column)
        table.ref = f"A1:{col_letter}{ws.max_row}"

        # prevent huge file
        if ws.max_row > 5000:
            ws.delete_rows(2, 1000)

        wb.save(LOG_FILE)


# ================= SCAN LOOP =================
def scan_loop():
    print("[spotfutures] Starting Spot vs Futures Scanner...")

    while True:
        try:
            spot = get_spot_prices()
            futures = get_futures_prices()
            funding = get_funding_rates()

            all_spreads = get_all_spreads(spot, futures, funding)
            opps = [row for row in all_spreads if row["is_opportunity"]]

            excel_df = pd.DataFrame([
                {
                    "Time": opp["time"],
                    "Symbol": opp["symbol"],
                    "Spot": opp["spot"],
                    "Futures": opp["futures"],
                    "Spread (%)": opp["spread_percent"],
                    "Net (%)": opp["net_percent"],
                }
                for opp in opps
            ])

            update_live(excel_df)
            append_log_table(excel_df)

            if opps:
                print(f"[spotfutures] Top: {opps[0]}")
            else:
                print(f"[spotfutures] No opportunities ({len(all_spreads)} pairs scanned)")

            send_opportunity({
                "type": "status",
                "time": datetime.now().strftime("%H:%M:%S"),
                "pairs": all_spreads,
            })

        except Exception as e:
            print("[spotfutures] Error:", e)

        time.sleep(1)


# ================= STANDALONE RUN (local testing only) =================
# When imported by scanner.py, none of this executes — scanner.py wires
# scan_loop() and the websocket route into its own shared FastAPI app instead.
if __name__ == "__main__":
    import threading
    import uvicorn
    from fastapi import FastAPI, WebSocket, WebSocketDisconnect
    from fastapi.middleware.cors import CORSMiddleware

    app = FastAPI()

    app.add_middleware(
        CORSMiddleware,
        allow_origins=["http://localhost:3000"],
        allow_credentials=True,
        allow_methods=["*"],
        allow_headers=["*"],
    )

    @app.websocket("/ws/spotfutures")
    async def spotfutures_websocket(websocket: WebSocket):
        await websocket.accept()
        clients.append(websocket)
        print("[spotfutures] Website connected")

        try:
            while True:
                await websocket.receive_text()
        except WebSocketDisconnect:
            clients.remove(websocket)
            print("[spotfutures] Website disconnected")

    @app.on_event("startup")
    async def startup_event():
        global main_loop
        main_loop = asyncio.get_running_loop()

        threading.Thread(target=scan_loop, daemon=True).start()

    uvicorn.run(
        app,
        host="0.0.0.0",
        port=int(os.environ.get("PORT", 8001)),
    )
