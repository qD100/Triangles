import requests
import time
import pandas as pd
from datetime import datetime
import os
import asyncio
import json
import threading
import collections

import uvicorn

from fastapi import FastAPI, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

import sf
import options


BASE = "USDT"
MAX_PROFIT_SANITY_CAP = 0.5

# User-configurable via the frontend settings panel (sent over /ws/arbitrage).
# fee_percent: deductible exchange fee per leg, as a percent (0.10 = 0.10%).
# min_profit_percent: minimum profit after fees, as a percent (0.05 = 0.05%).
settings = {
    "fee_percent": 0.10,
    "min_profit_percent": 0.05,
}

# Persistent background-engine state — survives page refreshes and new
# connections because it lives in this process, not in any one browser tab.
# Only resets on an actual server restart.
SCANNER_STARTED_AT = time.time() * 1000  # ms epoch, matches JS Date.now()
EVENTS_HISTORY_LIMIT = 500

events_history = collections.deque(maxlen=EVENTS_HISTORY_LIMIT)
best_profit = 0.0

LIVE_FILE = "live.xlsx"
LOG_FILE = "log.xlsx"
TABLE_NAME = "ArbTable"


# ================= FASTAPI =================

app = FastAPI()

allowed_origins = [
    origin.strip()
    for origin in os.environ.get("ALLOWED_ORIGINS", "http://localhost:3000").split(",")
    if origin.strip()
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

clients = []
main_loop = None


@app.websocket("/ws/arbitrage")
async def arbitrage_websocket(websocket: WebSocket):
    await websocket.accept()

    clients.append(websocket)

    print("Website connected")

    # Snapshot of persistent engine state, sent once so a fresh page load
    # (or refresh) catches up instantly instead of starting from zero.
    await websocket.send_json({
        "type": "init",
        "started_at": SCANNER_STARTED_AT,
        # Newest first, matching the order the frontend already keeps its
        # live-appended event list in.
        "events": list(reversed(events_history)),
        "best_profit": best_profit,
    })

    await websocket.send_json({"type": "settings", **settings})

    try:
        while True:
            message = await websocket.receive_text()

            try:
                payload = json.loads(message)
            except ValueError:
                continue

            if payload.get("type") != "settings":
                continue

            if "fee_percent" in payload:
                settings["fee_percent"] = max(0, float(payload["fee_percent"]))

            if "min_profit_percent" in payload:
                settings["min_profit_percent"] = max(0, float(payload["min_profit_percent"]))

            await broadcast_opportunity({"type": "settings", **settings})

    except WebSocketDisconnect:
        clients.remove(websocket)
        print("Website disconnected")


@app.websocket("/ws/spotfutures")
async def spotfutures_websocket(websocket: WebSocket):
    await websocket.accept()

    sf.clients.append(websocket)

    print("[spotfutures] Website connected")

    await websocket.send_json({
        "type": "init",
        "started_at": sf.SPOTFUTURES_STARTED_AT,
    })

    try:
        while True:
            message = await websocket.receive_text()
            await sf.handle_client_message(websocket, message)

    except WebSocketDisconnect:
        sf.clients.remove(websocket)
        print("[spotfutures] Website disconnected")


@app.websocket("/ws/options")
async def options_websocket(websocket: WebSocket):
    await websocket.accept()

    options.clients.append(websocket)

    print("[options] Website connected")

    await websocket.send_json({
        "type": "init",
        "started_at": options.OPTIONS_STARTED_AT,
        "events": list(options.events_history),
        "market": dict(options.market_state),
        "scanners": {k: dict(v) for k, v in options.scanner_status.items()},
        "settings": options.settings,
        "paper_trading": options.paper_trading_snapshot(),
    })

    try:
        while True:
            message = await websocket.receive_text()
            await options.handle_client_message(websocket, message)

    except WebSocketDisconnect:
        options.clients.remove(websocket)
        print("[options] Website disconnected")


async def broadcast_opportunity(opportunity):
    dead_clients = []

    for client in clients:
        try:
            await client.send_json(opportunity)

        except:
            dead_clients.append(client)

    for client in dead_clients:
        if client in clients:
            clients.remove(client)


def send_opportunity(opportunity):
    if main_loop is None:
        return

    asyncio.run_coroutine_threadsafe(
        broadcast_opportunity(opportunity),
        main_loop
    )


# ================= API =================

def get_exchange_info():
    url = "https://api.binance.com/api/v3/exchangeInfo"

    response = requests.get(url)
    data = response.json()

    if "symbols" not in data:
        raise RuntimeError(
            f"Binance exchangeInfo request failed "
            f"(HTTP {response.status_code}): {data}"
        )

    return [
        {
            "symbol": s["symbol"],
            "base": s["baseAsset"],
            "quote": s["quoteAsset"]
        }
        for s in data["symbols"]
        if s["status"] == "TRADING"
    ]


def get_prices():
    url = "https://api.binance.com/api/v3/ticker/bookTicker"

    data = requests.get(url).json()

    return {
        item["symbol"]: {
            "bid": float(item["bidPrice"]),
            "ask": float(item["askPrice"])
        }
        for item in data
    }


# ================= FILTERS =================

def is_spread_ok(prices, symbol, max_spread=0.002):
    bid = prices[symbol]["bid"]
    ask = prices[symbol]["ask"]

    return (
        ask > 0
        and (ask - bid) / ask < max_spread
    )


def get_top_usdt_coins(prices, symbols, top_n=25):
    liquidity = {}

    for s in symbols:
        if (
            s["quote"] == BASE
            and s["symbol"] in prices
        ):
            liquidity[s["base"]] = (
                prices[s["symbol"]]["bid"]
                * prices[s["symbol"]]["ask"]
            )

    sorted_coins = sorted(
        liquidity.items(),
        key=lambda x: -x[1]
    )

    return set(
        coin
        for coin, _ in sorted_coins[:top_n]
    )


# ================= ARBITRAGE =================

def find_triangles(symbols, prices, allowed):
    opps = []

    for s1 in symbols:
        if s1["quote"] != BASE:
            continue

        coin1 = s1["base"]
        sym1 = s1["symbol"]

        if (
            coin1 not in allowed
            or sym1 not in prices
            or not is_spread_ok(prices, sym1)
        ):
            continue

        for s2 in symbols:
            if s2["quote"] != coin1:
                continue

            coin2 = s2["base"]
            sym2 = s2["symbol"]

            if (
                coin2 not in allowed
                or sym2 not in prices
                or not is_spread_ok(prices, sym2)
            ):
                continue

            sym3 = coin2 + BASE

            if (
                sym3 not in prices
                or not is_spread_ok(prices, sym3)
            ):
                continue

            try:
                fee_fraction = settings["fee_percent"] / 100

                rate1 = 1 / prices[sym1]["ask"]
                rate2 = 1 / prices[sym2]["ask"]
                rate3 = prices[sym3]["bid"]

                final = (
                    rate1
                    * rate2
                    * rate3
                    * (1 - fee_fraction) ** 3
                )

                profit = (final - 1) * 100

                if settings["min_profit_percent"] < profit < MAX_PROFIT_SANITY_CAP:
                    time_str = datetime.now().strftime("%H:%M:%S")

                    opportunity = {
                        "type": "arbitrage",
                        "id": f"{time_str}-{coin1}-{coin2}",
                        "time": time_str,
                        "route": [
                            BASE,
                            coin1,
                            coin2,
                            BASE
                        ],
                        "path": (
                            f"{BASE} → "
                            f"{coin1} → "
                            f"{coin2} → "
                            f"{BASE}"
                        ),
                        "profit": round(profit, 5),
                        "live": True
                    }

                    opps.append(opportunity)

            except:
                continue

    return sorted(
        opps,
        key=lambda x: -x["profit"]
    )


# ================= EXCEL =================

def safe_write_live(df):
    try:
        df.to_excel(
            LIVE_FILE,
            index=False
        )

    except:
        pass


def append_log_table(df):
    if df.empty:
        return

    if not os.path.exists(LOG_FILE):
        wb = Workbook()

        ws = wb.active
        ws.title = "Log"

        headers = list(df.columns)

        ws.append(headers)

        for row in df.values:
            ws.append(list(row))

        end_row = ws.max_row
        end_col = ws.max_column

        col_letter = get_column_letter(end_col)

        table = Table(
            displayName=TABLE_NAME,
            ref=f"A1:{col_letter}{end_row}"
        )

        style = TableStyleInfo(
            name="TableStyleMedium9",
            showRowStripes=True
        )

        table.tableStyleInfo = style

        ws.add_table(table)

        wb.save(LOG_FILE)

    else:
        wb = load_workbook(LOG_FILE)

        ws = wb["Log"]

        for row in df.values:
            ws.append(list(row))

        table = ws.tables[TABLE_NAME]

        end_row = ws.max_row
        end_col = ws.max_column

        col_letter = get_column_letter(end_col)

        table.ref = (
            f"A1:{col_letter}{end_row}"
        )

        if ws.max_row > 5000:
            ws.delete_rows(2, 1000)

        wb.save(LOG_FILE)


# ================= SCANNER =================

def scanner():
    print("Loading Binance symbols...")

    symbols = None

    while symbols is None:
        try:
            symbols = get_exchange_info()
        except Exception as error:
            print("Failed to load Binance symbols, retrying in 5s:", error)
            time.sleep(5)

    print(
        f"{len(symbols)} symbols loaded"
    )

    while True:
        try:
            prices = get_prices()

            allowed = get_top_usdt_coins(
                prices,
                symbols
            )

            opps = find_triangles(
                symbols,
                prices,
                allowed
            )

            excel_opps = [
                {
                    "Time": opp["time"],
                    "Path": opp["path"],
                    "Profit (%)": opp["profit"]
                }
                for opp in opps
            ]

            df = pd.DataFrame(excel_opps)

            if not df.empty:
                live_df = df.head(10)

            else:
                live_df = pd.DataFrame([
                    {
                        "Time": datetime.now().strftime(
                            "%H:%M:%S"
                        ),
                        "Path": "No opportunities",
                        "Profit (%)": 0
                    }
                ])

            safe_write_live(live_df)

            append_log_table(df)

            print(
                f"{len(opps)} opportunities found"
            )

            # Send scan progress to website
            scanned_pairs = sum(1 for s in symbols if s["symbol"] in prices)

            send_opportunity({
                "type": "status",
                "time": datetime.now().strftime("%H:%M:%S"),
                "scanned_pairs": scanned_pairs,
                "total_pairs": len(symbols),
                "fee_percent": settings["fee_percent"],
                "min_profit_percent": settings["min_profit_percent"],
            })

            # Send opportunities to website, and retain them server-side so
            # a client connecting later (or reconnecting after a refresh)
            # can catch up via the "init" snapshot instead of seeing nothing
            # until the next detection.
            global best_profit

            for opportunity in opps[:10]:
                events_history.append(opportunity)
                best_profit = max(best_profit, opportunity["profit"])
                send_opportunity(opportunity)

        except Exception as error:
            print(
                "Scanner error:",
                error
            )

        time.sleep(1)


# ================= START =================

@app.on_event("startup")
async def startup_event():
    global main_loop

    main_loop = asyncio.get_running_loop()
    sf.main_loop = main_loop
    options.main_loop = main_loop

    scanner_thread = threading.Thread(
        target=scanner,
        daemon=True
    )

    scanner_thread.start()

    spotfutures_thread = threading.Thread(
        target=sf.scan_loop,
        daemon=True
    )

    spotfutures_thread.start()

    options_thread = threading.Thread(
        target=options.run_forever,
        daemon=True
    )

    options_thread.start()


if __name__ == "__main__":
    uvicorn.run(
        "scanner:app",
        host="0.0.0.0",
        port=int(os.environ.get("PORT", 8000)),
        reload=False,
    )