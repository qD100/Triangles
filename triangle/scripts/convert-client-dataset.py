"""One-time conversion of the client Excel dataset into clean JSON.

Reads app/namaa/mrktdata/<workbook>.xlsx and writes
app/namaa/mrktdata/clients/{clients,transactions}.json.

Deliberately never opens the "login credentials" sheet (plaintext
passwords / national-ID-like numbers) or the "لوحة العميل الفردية"
dashboard sheet (a formula template, not source data).

Run manually: python scripts/convert-client-dataset.py
"""

import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "app" / "namaa" / "mrktdata" / "نسخة نماء Banking Dataset Fanail.xlsx"
OUT_DIR = ROOT / "app" / "namaa" / "mrktdata" / "clients"

GENDER = {"ذكر": "male", "أنثى": "female"}

GOAL = {
    "شراء منزل": ("buy_home", "Buy a Home"),
    "بناء صندوق طوارئ": ("build_emergency_fund", "Build an Emergency Fund"),
    "التقاعد المبكر": ("early_retirement", "Early Retirement"),
    "تعليم الأبناء": ("children_education", "Children's Education"),
    "سداد الديون": ("debt_payoff", "Pay Off Debt"),
}

BEHAVIOR = {
    "عادي/متوازن": ("balanced", "Normal / Balanced"),
    "مدخر منضبط": ("disciplined_saver", "Disciplined Saver"),
    "دخل حر متذبذب": ("volatile_income", "Volatile Freelance Income"),
    "معتمد على القرض لتغطية اليومي": ("loan_dependent", "Relies on Loans for Daily Expenses"),
}

CITY = {
    "الرياض": "Riyadh",
    "جدة": "Jeddah",
    "الدمام": "Dammam",
    "مكة المكرمة": "Mecca",
}

INCOME_SOURCE = {
    "موظف حكومي": "Government Employee",
    "موظف قطاع خاص": "Private Sector Employee",
    "تجارة": "Business / Trade",
    "موظف شركة كبرى": "Large Company Employee",
}

MARITAL = {
    "متزوج مع أطفال": "Married with Children",
    "أعزب": "Single",
}

TX_TYPE = {"إيداع": "deposit", "سحب": "withdrawal"}

CATEGORY = {
    "إيجار": "rent",
    "اتصالات": "telecom",
    "ادخار": "savings",
    "استثمار": "investment",
    "اشتراكات": "subscriptions",
    "تحويل": "transfer",
    "ترفيه": "entertainment",
    "تسوق": "shopping",
    "تعليم": "education",
    "راتب": "salary",
    "سحب نقدي": "cash_withdrawal",
    "سفر": "travel",
    "صحة": "health",
    "صدقة": "charity",
    "صيدلية": "pharmacy",
    "قروض": "loans",
    "قهوة": "coffee",
    "مطاعم": "restaurants",
    "مواد غذائية": "groceries",
    "هدايا": "gifts",
    "وقود": "fuel",
}

SUPER_CATEGORY = {
    "احتياجات أساسية": "basic_needs",
    "محايد (دخل)": "neutral_income",
    "كماليات": "luxuries",
    "التزام مالي": "financial_obligation",
    "استثمار وادخار": "investment_savings",
    "نماء وعطاء": "growth_giving",
}

PAYMENT_METHOD = {
    "نقدي": "Cash",
    "بطاقة مدى": "Mada Card",
    "بطاقة ائتمانية": "Credit Card",
    "تحويل بنكي": "Bank Transfer",
    "STC Pay": "STC Pay",
    "Apple Pay": "Apple Pay",
}


def lookup(table, key, context):
    if key not in table:
        raise ValueError(f"Untranslated {context} value: {key!r}")
    return table[key]


def normalize_date(cell) -> str:
    if hasattr(cell, "date"):
        return cell.date().isoformat()
    s = str(cell).strip()
    if not re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        raise ValueError(f"Unrecognized date format: {s!r}")
    return s


def round2(x):
    return round(float(x), 2)


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)

    clients = []
    for row in wb["بيانات العملاء"].iter_rows(min_row=3, max_row=7, values_only=True):
        (
            cid, name, age, gender, city, salary, income_src, goal, balance,
            debt, installment, savings_goal, ef, marital, behavior,
        ) = row
        goal_slug, goal_label = lookup(GOAL, goal, "financial goal")
        beh_slug, beh_label = lookup(BEHAVIOR, behavior, "behavior pattern")
        clients.append({
            "id": cid,
            "fullName": name,
            "age": int(age),
            "gender": lookup(GENDER, gender, "gender"),
            "city": lookup(CITY, city, "city"),
            "baseSalary": round2(salary),
            "incomeSource": lookup(INCOME_SOURCE, income_src, "income source"),
            "financialGoal": goal_slug,
            "financialGoalLabel": goal_label,
            "currentBalance": round2(balance),
            "totalDebt": round2(debt),
            "monthlyInstallment": round2(installment),
            "monthlySavingsGoal": round2(savings_goal),
            "emergencyFund": round2(ef),
            "maritalStatus": lookup(MARITAL, marital, "marital status"),
            "behaviorPattern": beh_slug,
            "behaviorPatternLabel": beh_label,
        })

    transactions = []
    for row in wb["المعاملات المالية"].iter_rows(min_row=3, values_only=True):
        if row[0] is None:
            continue
        (tid, cid, date, time, merchant, typ, cat, amount, method, recurring, city, supercat) = row
        transactions.append({
            "id": tid,
            "clientId": cid,
            "date": normalize_date(date),
            "time": str(time),
            "merchant": merchant,
            "type": lookup(TX_TYPE, typ, "transaction type"),
            "category": lookup(CATEGORY, cat, "category"),
            "amount": round2(amount),
            "paymentMethod": lookup(PAYMENT_METHOD, method, "payment method"),
            "recurring": recurring == "نعم",
            "city": lookup(CITY, city, "city"),
            "superCategory": lookup(SUPER_CATEGORY, supercat, "super-category"),
        })

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / "clients.json").write_text(
        json.dumps(clients, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (OUT_DIR / "transactions.json").write_text(
        json.dumps(transactions, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"Wrote {len(clients)} clients, {len(transactions)} transactions to {OUT_DIR}")


if __name__ == "__main__":
    main()
